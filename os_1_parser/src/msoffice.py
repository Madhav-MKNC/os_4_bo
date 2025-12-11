import pandas as pd
from openpyxl.styles import PatternFill
from docx import Document

from src.address import Address
from src.colors import *

class MsOffice:
    def __init__(self):
        self.record_per_sheet = 64_000

    def export_to_MS_word(self, address_list: list, file_name: str):
        document = Document()
        table = document.add_table(rows=len(address_list), cols=3)
        table.style = 'Table Grid'
        for i, address in enumerate(address_list):
            row = table.rows[i // 3]  # Divide into 3 columns
            cell = row.cells[i % 3]
            cell.text = address.address
        document.save(file_name)

    def export_to_MS_Excel(self, address_list: list, file_name: str):
        print(f"{YELLOW}[ Exporting {len(address_list)} addresses to Excel file: {file_name} ]{RESET}")

        # Prepare data for DataFrame
        data = [[
            address.address_old, address.address, address.state, address.district,
            address.block, address.pin, address.country_code,
            address.phone, address.alternate_phone, 
            "YES" if address.is_reorder else "NO",
            address.name, address.district_from_address, address.state_from_address,
            address.occ_count, address.dist_matches_pin_and_addr, address.state_matches_pin_and_addr,
            address.book_name, address.book_lang, "YES" if address.is_repeat else "NO", address.email, address.faulty
        ] for address in address_list]
        
        # Create DataFrame for easier handling
        columns = [
            "ADDRESS ORIGINAL", "ADDRESS UPDATED", "STATE", "DISTRICT", "BLOCK", "PIN", 
            "COUNTRY CODE", "PHONE", "ALTERNATE PHONE", "RE_ORDER",
            "NAME", "DISTRICT_FROM_ADDRESS", "STATE_FROM_ADDRESS", "DISTRICT_MATCH_COUNT",
            "DIST_MATCHES_PIN_AND_ADDR", "STATE_MATCHES_PIN_AND_ADDR", "BOOK NAME", "BOOK LANG", 
            "REPEAT ORDER", "EMAIL", "FAULTY"
        ]
        
        df = pd.DataFrame(data, columns=columns)

        # Write to Excel with conditional formatting
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Addresses")
            workbook = writer.book
            worksheet = workbook["Addresses"]

            # Defining styles for the conditional formatting
            style_faulty = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # Pink
            style_warn = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            style_alert = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
            style_repeat = PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")  # Brown
            style_duplicate = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Gray

            # Apply color formatting based on conditions
            try:
                for row_idx, address in enumerate(address_list, start=2):  # start=2 to skip header row
                    row = worksheet[row_idx]
                    
                    if not address.phone: address.faulty = "FAULTY"

                    if address.faulty:
                        # faulty data
                        # Testing: Checking all faulty rows
                        with open('tmp/faulty_rows.txt', 'a', encoding='utf-8') as file:
                            file.write(f"\n{address.address_old}\n{address.address}\n")

                    if address.faulty and address.pin is None:
                        for col_idx in range(len(row)):
                            row[col_idx].fill = style_alert

                    elif address.faulty and address.phone is None:
                        for col_idx in range(len(row)):
                            row[col_idx].fill = style_faulty

                    elif address.is_repeat:
                        # Apply "repeat" color
                        for col_idx in range(len(row)):
                            row[col_idx].fill = style_repeat

                    elif address.is_reorder:
                        # Apply "duplicate" color
                        for col_idx in range(len(row)):
                            row[col_idx].fill = style_duplicate

                    elif address.pin is not None and address.phone is not None:
                        # No specific color; regular address
                        pass

                    elif address.phone is not None:
                        # Apply "alert" color
                        for col_idx in range(len(row)):
                            row[col_idx].fill = style_alert

                    else:
                        # Apply "warn" color
                        for col_idx in range(len(row)):
                            row[col_idx].fill = style_warn

                    # Print progress bar (every 1000 rows, adjust as needed)
                    if row_idx % 10 == 0 or row_idx == len(address_list):
                        progress = (row_idx / len(address_list)) * 100
                        bar = '█' * int(progress // 2) + '-' * (50 - int(progress // 2))
                        color = GREEN if progress == 100 else YELLOW
                        print(f"\r{color}[{bar}] {progress:.2f}% Complete{RESET}", end='', flush=True)

            except Exception as e:
                print(f"\n{RED}[ERROR] An error occurred during formatting: {e}{RESET}")

            # Set column widths based on header length
            for col in worksheet.columns:
                max_length = len(str(col[0].value))  # Get the header length only
                column = col[0].column_letter  # Get the column name
                adjusted_width = (max_length + 2)  # Add some padding to ensure visibility
                worksheet.column_dimensions[column].width = adjusted_width
        
        print(f"\n{GREEN}[ Export completed: {file_name} ✔ ]{RESET}")

    def export_to_MS_Excel_using_xlsxwriter(self, address_list: list, file_name: str):
        print(f"{YELLOW}[ Exporting {len(address_list)} addresses to Excel file: {file_name} ]{RESET}")

        # Build DataFrame (vectorized; no per-cell ops)
        data = [[
            address.address_old, address.address, address.state, address.district,
            address.block, address.pin, address.country_code,
            address.phone, address.alternate_phone, 
            "YES" if address.is_reorder else "NO",
            address.name, address.district_from_address, address.state_from_address,
            address.occ_count, address.dist_matches_pin_and_addr, address.state_matches_pin_and_addr,
            address.book_name, address.book_lang, "YES" if address.is_repeat else "NO", address.email, address.faulty
        ] for address in address_list]

        columns = [
            "ADDRESS ORIGINAL", "ADDRESS UPDATED", "STATE", "DISTRICT", "BLOCK", "PIN", 
            "COUNTRY CODE", "PHONE", "ALTERNATE PHONE", "RE_ORDER",
            "NAME", "DISTRICT_FROM_ADDRESS", "STATE_FROM_ADDRESS", "DISTRICT_MATCH_COUNT",
            "DIST_MATCHES_PIN_AND_ADDR", "STATE_MATCHES_PIN_AND_ADDR", "BOOK NAME", "BOOK LANG", 
            "REPEAT ORDER", "EMAIL", "FAULTY"
        ]

        df = pd.DataFrame(data, columns=columns)

        # Fast writer + conditional formatting (no Python loops over rows)
        with pd.ExcelWriter(file_name, engine="xlsxwriter", engine_kwargs={"options": {"strings_to_urls": False}}) as writer:
            sheet = "Addresses"
            df.to_excel(writer, index=False, sheet_name=sheet)
            wb  = writer.book
            ws  = writer.sheets[sheet]
            n   = len(df) + 1  # header at row 1, data starts at 2
            rng = f"A2:U{n}"

            fmt_faulty_pin   = wb.add_format({"bg_color": "#FFFF00"})  # Yellow
            fmt_faulty_phone = wb.add_format({"bg_color": "#FFC0CB"})  # Pink
            fmt_repeat       = wb.add_format({"bg_color": "#A52A2A"})  # Brown
            fmt_duplicate    = wb.add_format({"bg_color": "#808080"})  # Gray
            fmt_warn         = wb.add_format({"bg_color": "#FF0000"})  # Red
            fmt_alert        = wb.add_format({"bg_color": "#FFFF00"})  # Yellow

            # Column map (A=1): F=PIN, H=PHONE, J=RE_ORDER, S=REPEAT ORDER, U=FAULTY
            # Priority top→down; stop_if_true keeps later rules from repainting
            ws.conditional_format(rng, {
                "type": "formula", "criteria": '=$S2="YES"', "format": fmt_repeat, "stop_if_true": True
            })
            ws.conditional_format(rng, {
                "type": "formula", "criteria": '=$J2="YES"', "format": fmt_duplicate, "stop_if_true": True
            })
            ws.conditional_format(rng, {
                "type": "formula", "criteria": '=AND($U2<>"",$F2="")', "format": fmt_faulty_pin, "stop_if_true": True
            })
            ws.conditional_format(rng, {
                "type": "formula", "criteria": '=AND($U2<>"",$H2="")', "format": fmt_faulty_phone, "stop_if_true": True
            })
            ws.conditional_format(rng, {
                "type": "formula", "criteria": '=AND($U2="", $H2<>"", $F2="")', "format": fmt_alert, "stop_if_true": True
            })
            ws.conditional_format(rng, {
                "type": "formula", "criteria": '=AND($U2="", $H2="")', "format": fmt_warn, "stop_if_true": True
            })

            # Reasonable fixed widths (fast). Tweak as needed.
            widths = {
                "A": 45, "B": 45, "C": 14, "D": 18, "E": 18, "F": 10, "G": 12, "H": 14, "I": 16, "J": 10,
                "K": 18, "L": 22, "M": 22, "N": 10, "O": 14, "P": 16, "Q": 18, "R": 14, "S": 12, "T": 28, "U": 10
            }
            for col, w in widths.items():
                ws.set_column(f"{col}:{col}", w)

        print(f"{GREEN}[ Export completed: {file_name} ✔ ]{RESET}")

    def import_from_Excel_sheet(self, file_name: str):
        # Read the Excel file into a DataFrame
        df = pd.read_excel(file_name)

        # Convert DataFrame rows into Address objects
        address_list = []
        for _, row in df.iterrows():
            address = Address(
                address_old=row["ADDRESS ORIGINAL"],
                address=row["ADDRESS UPDATED"],
                state=row["STATE"] if pd.notna(row["STATE"]) else None,
                district=row["DISTRICT"] if pd.notna(row["DISTRICT"]) else None,
                block=row["BLOCK"] if pd.notna(row["BLOCK"]) else None,
                pin=row["PIN"] if pd.notna(row["PIN"]) else None,
                country_code=row["COUNTRY CODE"],
                phone=row["PHONE"] if pd.notna(row["PHONE"]) else None,
                alternate_phone=row["ALTERNATE PHONE"],
                is_reorder=True if row["RE_ORDER"] == "YES" else False,
                name=row["NAME"],
                district_from_address=row["DISTRICT_FROM_ADDRESS"] if pd.notna(row["DISTRICT_FROM_ADDRESS"]) else None,
                state_from_address=row["STATE_FROM_ADDRESS"] if pd.notna(row["STATE_FROM_ADDRESS"]) else None,
                occ_count=row["DISTRICT_MATCH_COUNT"],
                dist_matches_pin_and_addr=row["DIST_MATCHES_PIN_AND_ADDR"],
                state_matches_pin_and_addr=row["STATE_MATCHES_PIN_AND_ADDR"],
                book_name=row["BOOK NAME"],
                book_lang=row["BOOK LANG"],
                is_repeat=True if row["REPEAT ORDER"] == "YES" else False,
                email=row["EMAIL"],
                faulty=row["FAULTY"]
            )
            address_list.append(address)
        return address_list
